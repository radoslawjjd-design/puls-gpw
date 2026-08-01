"""The only place in the broker package that touches spreadsheet bytes.

Isolated so the parsers can be tested against plain dicts. Mirrors how
``scripts/backfill_historical_closes.py`` keeps its pure logic away from I/O.
"""

import io
import zipfile

from openpyxl import load_workbook

from src.brokers.errors import BrokerFileTooLargeError, BrokerParseError

# Column labels a sheet must carry. Rows are keyed by header, never by position:
# positional indexing is what let a defect live for a month in PUL-98.
_REQUIRED_COLUMNS = {
    "Cash Operations": ("Type", "Ticker", "Instrument", "Time", "Amount", "ID", "Comment"),
}

# How far down to look for the header. XTB puts an account-number/date-range
# preamble above it, so the header row is located by its labels rather than
# assumed to sit at a fixed index.
_MAX_HEADER_SCAN = 25

# Ceiling on the workbook's decompressed size, checked before openpyxl opens it.
# The upload gate in the API measures compressed bytes, which says nothing about
# what a zip expands to: a measured 1 MB upload of repetitive shared strings
# reaches 42 MB decompressed and 85 MB of peak heap. This has to run before
# ``load_workbook`` because that call parses the whole shared-string table
# eagerly, even in read-only mode, before any row is iterated.
#
# 8 MiB is roughly 200x the largest real XTB export (41 KB), against a 512 MiB
# Cloud Run instance shared by concurrent requests.
_MAX_UNCOMPRESSED_BYTES = 8 * 1024 * 1024

# Cells the reader will materialise. The byte ceiling bounds the input; this
# bounds what iterating it produces, which is the larger number: a cell costs
# far more as a Python object than as XML. ~50x the cell count of the largest
# real export.
_MAX_CELLS = 200_000

# Parts the archive may declare. The byte ceiling below is blind to entry COUNT:
# ~97k zero-length entries fit inside the 5 MB upload gate and cost ~47 MiB of
# ZipInfo objects while summing to zero bytes. A real .xlsx has well under a
# hundred parts.
_MAX_ZIP_ENTRIES = 256


def read_sheets(data: bytes) -> dict[str, list[dict]]:
    """Read a workbook into ``{sheet name: [row dict, ...]}``.

    Only sheets listed in ``_REQUIRED_COLUMNS`` are read; anything else in the
    file is ignored, so a caller asking for a sheet that is absent from the
    result cannot tell "not in the file" from "not one we read" — and does not
    need to, since a parser only ever asks for sheets it declared.

    Rows are keyed by column header. A sheet missing an expected column raises
    ``BrokerParseError`` naming that column rather than guessing.
    """
    _reject_oversized(data)
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except BrokerParseError:
        raise
    except Exception as exc:  # openpyxl raises a wide range for malformed input
        raise BrokerParseError(f"nie udalo sie odczytac pliku xlsx: {exc}") from exc

    budget = _CellBudget(_MAX_CELLS)
    try:
        # Only the sheets a parser declares columns for. A real export carries
        # several; the XTB parser indexes exactly one. Reading the rest cost
        # memory on every honest import and gave a payload somewhere to hide.
        # This deliberately couples "what we read" to "what we validate" — if a
        # broker ever needs a sheet with no required columns, that is the moment
        # to split the two, not before.
        return {
            sheet.title: _read_one(sheet, budget)
            for sheet in workbook.worksheets
            if sheet.title in _REQUIRED_COLUMNS
        }
    finally:
        workbook.close()


class _CellBudget:
    """Cells the whole workbook may materialise — not each sheet on its own.

    A payload spread across twenty sheets would clear any per-sheet cap while
    costing exactly the same memory, so the budget is shared and spent down.
    """

    def __init__(self, remaining: int) -> None:
        self.remaining = remaining

    def spend(self, cells: int) -> None:
        self.remaining -= cells
        if self.remaining < 0:
            raise BrokerFileTooLargeError(
                f"plik zawiera ponad {_MAX_CELLS - self.remaining} komorek do wczytania, "
                f"limit to {_MAX_CELLS}"
            )


def _reject_oversized(data: bytes) -> None:
    """Refuse a workbook whose parts add up to more than the ceiling.

    The sizes come from the zip central directory, which is attacker-controlled —
    but ``ZipFile.open`` caps every read at the declared ``file_size``, so a
    manifest that under-reports yields truncated data and a CRC failure rather
    than an unbounded read. That makes the sum a sound upper bound.
    """
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            entries = archive.infolist()
            if len(entries) > _MAX_ZIP_ENTRIES:
                raise BrokerFileTooLargeError(
                    f"plik zawiera {len(entries)} czesci, limit to {_MAX_ZIP_ENTRIES}"
                )
            total = sum(item.file_size for item in entries)
    except (zipfile.BadZipFile, NotImplementedError) as exc:
        # Opening the archive first moves these failures ahead of load_workbook,
        # so the message it used to produce has to be produced here instead.
        # NotImplementedError is not hypothetical: ZipFile raises it for an
        # extract_version above 63, which is one byte of the central directory.
        raise BrokerParseError(f"nie udalo sie odczytac pliku xlsx: {exc}") from exc

    if total > _MAX_UNCOMPRESSED_BYTES:
        megabytes = _MAX_UNCOMPRESSED_BYTES // (1024 * 1024)
        raise BrokerFileTooLargeError(
            f"plik po rozpakowaniu zajmuje {total / 1024 / 1024:.1f} MB, "
            f"limit to {megabytes} MB"
        )


def _read_one(sheet, budget: _CellBudget) -> list[dict]:
    # XTB declares the sheet dimension as a single cell no matter how much data
    # it holds. In read-only mode openpyxl trusts that declaration, so without
    # this reset the import comes back with one row instead of hundreds —
    # silently empty rather than loudly broken.
    if hasattr(sheet, "reset_dimensions"):
        sheet.reset_dimensions()

    # iter_rows is a generator that reads the sheet XML incrementally; the list()
    # this replaces is what threw that property away. Spending the budget as the
    # rows arrive means an oversized sheet is abandoned partway rather than
    # materialised and then judged.
    rows: list[tuple] = []
    for row in sheet.iter_rows(values_only=True):
        # max(..., 1): openpyxl fills gaps in row numbering with empty tuples,
        # and after reset_dimensions() those are length zero. Charging only for
        # cells lets a single `<row r="20000000">` — four kilobytes of file —
        # materialise twenty million list entries for free, spending nothing.
        # Charging for the row itself binds their count to the same budget.
        budget.spend(max(len(row), 1))
        rows.append(row)
    # Guaranteed present: read_sheets only hands over sheets it looked up here.
    required = _REQUIRED_COLUMNS[sheet.title]
    header_index = _find_header(rows, required)
    if header_index is None:
        missing = _missing_columns(rows, required)
        raise BrokerParseError(
            f"arkusz '{sheet.title}' nie ma oczekiwanych kolumn: {', '.join(missing)}"
        )

    header = [str(cell).strip() if cell is not None else "" for cell in rows[header_index]]
    # Dropped in place rather than sliced: a slice would hold a second near-full
    # copy of the workbook alive, doubling peak memory at exactly the moment the
    # budget has just declared this file acceptable.
    del rows[:header_index + 1]
    out: list[dict] = []
    for row in rows:
        if row is None or all(cell is None or cell == "" for cell in row):
            continue
        out.append({name: value for name, value in zip(header, row) if name})
    return out


def _find_header(rows: list[tuple], required: tuple[str, ...]) -> int | None:
    wanted = set(required)
    for index, row in enumerate(rows[:_MAX_HEADER_SCAN]):
        labels = {str(cell).strip() for cell in row if cell is not None}
        if wanted <= labels:
            return index
    return None


def _missing_columns(rows: list[tuple], required: tuple[str, ...]) -> list[str]:
    """Best-effort report of which expected columns were absent."""
    best: list[str] = list(required)
    for row in rows[:_MAX_HEADER_SCAN]:
        labels = {str(cell).strip() for cell in row if cell is not None}
        missing = [name for name in required if name not in labels]
        if len(missing) < len(best):
            best = missing
    return best
