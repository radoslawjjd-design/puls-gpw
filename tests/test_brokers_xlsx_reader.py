import io
import zipfile

import pytest
from openpyxl import Workbook

from src.brokers import xlsx_reader
from src.brokers.errors import BrokerFileTooLargeError, BrokerParseError
from src.brokers.xlsx_reader import read_sheets

_HEADER = ["Type", "Ticker", "Instrument", "Time", "Amount", "ID", "Comment", "Product"]

_SHARED_STRINGS_CT = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"
)


def _workbook_bytes(*, preamble_rows: int = 4, header: list | None = None,
                    data: list[list] | None = None, broken_dimension: bool = False,
                    shared_strings_bytes: int = 0,
                    extra_sheets: dict[str, list[list]] | None = None) -> bytes:
    """Build an XTB-shaped workbook in memory.

    The real exports must never be committed (they carry account numbers), so
    every fixture is synthesized to the same shape instead.

    ``shared_strings_bytes`` injects a shared-string table of roughly that many
    decompressed bytes. openpyxl writes inline strings and emits no such part on
    its own, so it is grafted in along with its manifest entry — which is how
    openpyxl finds it (``reader/excel.py``, ``read_strings``). The worksheet
    never references those strings: the point is that they are loaded eagerly by
    ``load_workbook`` regardless, so nothing inside the row iteration can see
    them coming.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Operations"
    for i in range(preamble_rows):
        ws.append([f"Meta {i}", "value"])
    ws.append(header if header is not None else _HEADER)
    for row in data or []:
        ws.append(row)
    if broken_dimension:
        # XTB declares "A1:A1" regardless of content. openpyxl's read-only mode
        # trusts that declaration and yields a single row unless it is reset.
        ws.calculate_dimension = lambda **kwargs: "A1:A1"
    for title, rows in (extra_sheets or {}).items():
        extra = wb.create_sheet(title)
        for row in rows:
            extra.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    if not shared_strings_bytes:
        return buffer.getvalue()
    return _with_shared_strings(buffer.getvalue(), shared_strings_bytes)


def _sparse_workbook_bytes(last_row: int) -> bytes:
    """A workbook whose sheet declares one row far down an otherwise empty sheet.

    openpyxl's writer cannot emit a sparse ``r=``, so the sheet part is written
    by hand. Every index between the header and ``last_row`` is a gap openpyxl
    fills with an empty tuple — which is the whole point of the fixture.
    """
    cells = "".join(
        f'<c r="{chr(65 + i)}1" t="inlineStr"><is><t>{name}</t></is></c>'
        for i, name in enumerate(_HEADER)
    )
    sheet = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<dimension ref="A1:A1"/><sheetData>'
        f'<row r="1">{cells}</row>'
        f'<row r="{last_row}"><c r="A{last_row}" t="inlineStr"><is><t>x</t></is></c></row>'
        "</sheetData></worksheet>"
    ).encode()

    source = zipfile.ZipFile(io.BytesIO(_workbook_bytes()))
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as target:
        for item in source.infolist():
            body = sheet if item.filename == "xl/worksheets/sheet1.xml" else source.read(item.filename)
            target.writestr(item.filename, body)
    source.close()
    return out.getvalue()


def _with_shared_strings(workbook: bytes, approx_bytes: int) -> bytes:
    entry = "<si><t>" + ("x" * 96) + "</t></si>"
    count = max(1, approx_bytes // len(entry))
    payload = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        f'count="{count}" uniqueCount="{count}">' + entry * count + "</sst>"
    ).encode()

    source = zipfile.ZipFile(io.BytesIO(workbook))
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as target:
        for item in source.infolist():
            body = source.read(item.filename)
            if item.filename == "[Content_Types].xml":
                body = body.decode().replace(
                    "</Types>",
                    f'<Override PartName="/xl/sharedStrings.xml" '
                    f'ContentType="{_SHARED_STRINGS_CT}"/></Types>',
                ).encode()
            target.writestr(item.filename, body)
        target.writestr("xl/sharedStrings.xml", payload)
    source.close()
    return out.getvalue()


def test_rows_are_keyed_by_header_never_by_position():
    data = [["Stock purchase", "TOA.PL", "Toya", None, -4128.24, "1", "OPEN BUY 412 @ 10.02", "My Trades"]]

    sheets = read_sheets(_workbook_bytes(data=data))

    assert sheets["Cash Operations"][0]["Ticker"] == "TOA.PL"
    assert sheets["Cash Operations"][0]["Amount"] == pytest.approx(-4128.24)


def test_header_is_found_below_the_metadata_preamble():
    # XTB puts account number and date range above the header; the header is not
    # at a fixed row, so it is located by its labels.
    data = [["Dividend", "KRU.PL", "Kruk", None, 722.0, "1", "", "My Trades"]]

    sheets = read_sheets(_workbook_bytes(preamble_rows=7, data=data))

    assert len(sheets["Cash Operations"]) == 1
    assert sheets["Cash Operations"][0]["Type"] == "Dividend"


def test_a_missing_expected_column_raises_naming_it():
    header = [c for c in _HEADER if c != "Amount"]

    with pytest.raises(BrokerParseError) as excinfo:
        read_sheets(_workbook_bytes(header=header))

    assert "Amount" in str(excinfo.value)


def test_all_rows_are_read_even_when_the_file_declares_a_1x1_dimension():
    # This is the real trap: XTB's export declares its dimension as a single
    # cell. Reading it in openpyxl's read-only mode without resetting that
    # declaration yields ONE row, so the import silently comes back empty
    # instead of failing loudly.
    data = [
        ["Dividend", "KRU.PL", "Kruk", None, float(n), str(n), "", "My Trades"]
        for n in range(1, 26)
    ]

    sheets = read_sheets(_workbook_bytes(data=data, broken_dimension=True))

    assert len(sheets["Cash Operations"]) == 25


def test_a_file_that_is_not_a_workbook_raises_a_parse_error():
    with pytest.raises(BrokerParseError):
        read_sheets(b"this is not a spreadsheet")


# ── size ceiling (PUL-105) ──────────────────────────────────────────────


def test_a_workbook_that_decompresses_past_the_ceiling_is_refused():
    # A megabyte of upload expanding to tens of megabytes of shared strings sits
    # comfortably under the 5 MB gate on the compressed body. openpyxl loads that
    # table in full inside load_workbook, before a single row is iterated, so the
    # only place this can be stopped is before the workbook is opened at all.
    payload = _workbook_bytes(shared_strings_bytes=16 * 1024 * 1024)

    with pytest.raises(BrokerFileTooLargeError) as excinfo:
        read_sheets(payload)

    assert "limit to 8 MB" in str(excinfo.value)


def test_iteration_stops_once_the_cell_budget_is_spent(monkeypatch):
    # The byte ceiling bounds the input; this bounds what iterating it turns the
    # input into, where a cell costs far more as a Python object than as XML.
    # The budget is patched down rather than fabricating a 200k-cell fixture:
    # what is under test is that iteration aborts when the budget runs out, not
    # the value of the constant. Actual memory behaviour is verified manually —
    # see Phase 1 of the plan, and lessons.md on the limits of mocked tests.
    monkeypatch.setattr(xlsx_reader, "_MAX_CELLS", 40)
    data = [
        ["Dividend", "KRU.PL", "Kruk", None, float(n), str(n), "", "My Trades"]
        for n in range(1, 11)
    ]

    with pytest.raises(BrokerFileTooLargeError):
        read_sheets(_workbook_bytes(data=data))


def test_an_archive_with_absurdly_many_parts_is_refused():
    # Impl-review F3. The byte ceiling is blind to entry count: zero-length
    # entries sum to nothing while each one still costs a ZipInfo object.
    base = _workbook_bytes()
    out = io.BytesIO()
    with (
        zipfile.ZipFile(io.BytesIO(base)) as source,
        zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as target,
    ):
        for item in source.infolist():
            target.writestr(item.filename, source.read(item.filename))
        for i in range(300):
            target.writestr(f"xl/pad{i}.bin", b"")

    with pytest.raises(BrokerFileTooLargeError) as excinfo:
        read_sheets(out.getvalue())

    assert "czesci" in str(excinfo.value)


def test_an_archive_declaring_an_impossible_extract_version_is_a_parse_error():
    # Impl-review F2. Opening the zip ahead of load_workbook moved this failure
    # earlier: ZipFile raises NotImplementedError (not BadZipFile) for an
    # extract_version above 63 — one byte of the central directory. Before the
    # fix it escaped read_sheets entirely and the endpoint answered 500.
    raw = bytearray(_workbook_bytes())
    marker = raw.index(b"PK\x01\x02")
    raw[marker + 6:marker + 8] = (99).to_bytes(2, "little")

    with pytest.raises(BrokerParseError) as excinfo:
        read_sheets(bytes(raw))

    assert "nie udalo sie odczytac pliku xlsx" in str(excinfo.value)


def test_a_sheet_that_declares_one_row_far_down_is_refused():
    # Impl-review F1. Gaps in row numbering come back as EMPTY tuples once
    # reset_dimensions() has cleared max_column, so a budget that charges per
    # cell charges nothing for them — while the list of rows grows all the same.
    # The row number is attacker-supplied and unbounded, so four kilobytes of
    # file bought 305 MB of heap and still parsed "successfully".
    # Deliberately not monkeypatched: this must hold at the real ceiling.
    payload = _sparse_workbook_bytes(20_000_000)

    with pytest.raises(BrokerFileTooLargeError):
        read_sheets(payload)


def test_a_sheet_no_parser_asks_for_is_never_read(monkeypatch):
    # A real export carries several sheets; the XTB parser indexes exactly one.
    # Reading the rest cost memory on every honest import and handed an attacker
    # somewhere to hide a payload. Spending the cell budget is the observable
    # proof: with the budget too small for the second sheet, the read still
    # succeeds — because that sheet is never touched.
    monkeypatch.setattr(xlsx_reader, "_MAX_CELLS", 200)
    data = [["Dividend", "KRU.PL", "Kruk", None, 722.0, "1", "", "My Trades"]]
    bulky = [[f"cell {i}-{j}" for j in range(8)] for i in range(100)]

    sheets = read_sheets(_workbook_bytes(data=data, extra_sheets={"Open Positions": bulky}))

    assert list(sheets) == ["Cash Operations"]
    assert sheets["Cash Operations"][0]["Ticker"] == "KRU.PL"
