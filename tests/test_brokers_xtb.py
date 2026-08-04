import ast
import io
import pathlib
from dataclasses import replace
from datetime import datetime

import pytest
from openpyxl import Workbook

from src.brokers import get_parser, list_brokers
from src.brokers.errors import BrokerParseError, UnknownBrokerError
from src.brokers.xtb import (
    OP_BUY,
    OP_CASH,
    OP_DIVIDEND,
    OP_SELL,
    OP_WITHHOLDING_TAX,
    Operation,
    extract_cash_balance,
    extract_dividends,
    normalize_operations,
    parse_trade_comment,
    parse_xtb_export,
    reconstruct_positions,
)


def _row(type_, ticker="", instrument="", amount=0.0, oid="1", comment="", day=2, micro=0):
    return {
        "Type": type_,
        "Ticker": ticker,
        "Instrument": instrument,
        "Time": datetime(2026, 7, day, 9, 54, 1, micro),
        "Amount": amount,
        "ID": oid,
        "Comment": comment,
        "Product": "My Trades",
    }


def _trade(ticker: str, volume: float, price: float, day: int, op_type: str) -> Operation:
    return Operation(
        external_id=f"{op_type}-{ticker}-{day}-{volume}",
        op_type=op_type,
        raw_type="Stock purchase" if op_type == OP_BUY else "Stock sale",
        occurred_at=datetime(2025, 1, day, 10, 0),
        amount_pln=round(volume * price, 2) * (-1 if op_type == OP_BUY else 1),
        ticker=ticker,
        volume=volume,
        unit_price=price,
    )


def _buy(ticker: str, volume: float, price: float, day: int) -> Operation:
    return _trade(ticker, volume, price, day, OP_BUY)


def _sell(ticker: str, volume: float, price: float, day: int) -> Operation:
    return _trade(ticker, volume, price, day, OP_SELL)


def test_partial_fill_comment_yields_filled_volume_not_whole_order():
    # XTB splits one order into several fill rows. The number BEFORE the slash is
    # what THIS row filled; the one after it is the whole order. Reading the second
    # would double every volume in the reconstruction.
    assert parse_trade_comment("OPEN BUY 9/11 @ 188.40") == (9.0, 188.40)


def test_comment_without_slash_is_a_single_fill():
    assert parse_trade_comment("OPEN BUY 15 @ 103.80") == (15.0, 103.80)


def test_close_direction_and_fractional_whole_order():
    assert parse_trade_comment("CLOSE BUY 31/31.7931 @ 33.330") == (31.0, 33.330)


def test_fractional_filled_volume_is_preserved():
    assert parse_trade_comment("OPEN BUY 0.4962 @ 103.35") == (0.4962, 103.35)


def test_unparsable_comment_raises_instead_of_guessing():
    with pytest.raises(BrokerParseError):
        parse_trade_comment("free-form text that is not a trade")


def test_sale_consumes_the_oldest_lot_so_avg_is_not_the_weighted_mean():
    # The whole point of FIFO: after selling the cheap lot, the remaining cost
    # basis is the expensive one. A cost-weighted average over ALL buys would
    # say 150.00 — which is not what XTB shows in the app.
    positions, closed = reconstruct_positions([
        _buy("SNT", 10, 100.0, day=2),
        _buy("SNT", 10, 200.0, day=3),
        _sell("SNT", 10, 250.0, day=4),
    ])

    assert closed == []
    assert len(positions) == 1
    assert positions[0].shares == pytest.approx(10.0)
    assert positions[0].avg_buy_price == pytest.approx(200.0)


def test_partial_sale_cutting_into_a_lot_leaves_a_blended_basis():
    # Sell 5 — the first lot is only partly consumed, so 5 @ 100 survives
    # alongside 10 @ 200: remaining_cost 2500 / remaining_shares 15.
    positions, _ = reconstruct_positions([
        _buy("KRU", 10, 100.0, day=2),
        _buy("KRU", 10, 200.0, day=3),
        _sell("KRU", 5, 250.0, day=4),
    ])

    assert positions[0].shares == pytest.approx(15.0)
    assert positions[0].avg_buy_price == pytest.approx(2500.0 / 15.0)


def test_avg_buy_price_is_remaining_cost_over_remaining_shares():
    # This exact contract is load-bearing: _merge_positions_by_ticker averages
    # positions again by cost in "Wszystkie" mode, and three other places
    # compute P/L as (price - avg) * shares.
    positions, _ = reconstruct_positions([
        _buy("PAS", 30, 90.0, day=2),
        _buy("PAS", 20, 120.0, day=3),
        _sell("PAS", 10, 130.0, day=4),
    ])

    remaining_cost = 20 * 90.0 + 20 * 120.0
    assert positions[0].shares == pytest.approx(40.0)
    assert positions[0].avg_buy_price == pytest.approx(remaining_cost / 40.0)


def test_ticker_sold_to_zero_is_reported_closed_not_as_a_position():
    positions, closed = reconstruct_positions([
        _buy("PZU", 12, 50.0, day=2),
        _sell("PZU", 12, 55.0, day=5),
    ])

    assert positions == []
    assert closed == ["PZU"]


def test_fractional_shares_survive_the_round_trip():
    positions, _ = reconstruct_positions([
        _buy("ETFBW20TR", 10.2017, 52.89, day=2),
    ])

    assert positions[0].shares == pytest.approx(10.2017)
    assert positions[0].avg_buy_price == pytest.approx(52.89)


def test_pl_suffix_is_stripped_from_the_ticker():
    positions, closed = reconstruct_positions([
        _buy("TOA.PL", 412, 10.02, day=2),
        _buy("PZU.PL", 5, 50.0, day=2),
        _sell("PZU.PL", 5, 51.0, day=3),
    ])

    assert positions[0].ticker == "TOA"
    assert closed == ["PZU"]


def test_company_name_survives_a_first_row_that_carries_none():
    """Deliberate behaviour change (PUL-114). The old setdefault locked in None
    whenever the earliest row for a ticker happened to carry no instrument name,
    so that ticker lost its company name for good. The shared lot ledger keeps
    looking until it finds one."""
    first = _buy("DIG", 5, 60.0, day=2)
    later = _buy("DIG", 5, 70.0, day=3)
    positions, _ = reconstruct_positions([
        replace(first, instrument_name=None),
        replace(later, instrument_name="Digitree Group SA"),
    ])

    assert positions[0].company_name == "Digitree Group SA"


def test_withholding_tax_and_interest_tax_are_not_the_same_type():
    # Both are "tax" in plain English but only one belongs to the dividend
    # total. Lumping them together turns Glowny's -435.28 into -437.00.
    operations, _ = normalize_operations([
        _row("Dividend", "KRU.PL", "Kruk", amount=100.0, oid="a"),
        _row("Withholding tax", "KRU.PL", "Kruk", amount=-19.0, oid="b"),
        _row("Free funds interest tax", amount=-1.72, oid="c"),
    ])

    by_id = {op.external_id: op for op in operations}
    assert by_id["a"].op_type == OP_DIVIDEND
    assert by_id["b"].op_type == OP_WITHHOLDING_TAX
    assert by_id["c"].op_type == OP_CASH


def test_broker_labels_map_onto_normalized_types_keeping_the_original():
    operations, _ = normalize_operations([
        _row("Stock purchase", "TOA.PL", "Toya", amount=-4128.24, oid="a", comment="OPEN BUY 412 @ 10.02"),
        _row("Stock sell", "PZU.PL", "PZU", amount=550.0, oid="b", comment="CLOSE BUY 10 @ 55.00"),
        _row("IKZE deposit", amount=6473.97, oid="c"),
    ])

    by_id = {op.external_id: op for op in operations}
    assert by_id["a"].op_type == OP_BUY
    assert by_id["b"].op_type == OP_SELL
    assert by_id["c"].op_type == OP_CASH
    # The broker's own wording survives, so a future parser can map a different
    # vocabulary onto the same normalized set.
    assert by_id["a"].raw_type == "Stock purchase"


def test_trade_rows_get_volume_and_price_from_the_comment():
    operations, _ = normalize_operations([
        _row("Stock purchase", "CBF.PL", "CyberFolks", amount=-1695.6, oid="a",
             comment="OPEN BUY 9/11 @ 188.40"),
    ])

    assert operations[0].volume == pytest.approx(9.0)
    assert operations[0].unit_price == pytest.approx(188.40)


def test_total_row_is_dropped():
    operations, _ = normalize_operations([
        _row("Dividend", "KRU.PL", "Kruk", amount=100.0, oid="a"),
        _row("Total", amount=99999.0, oid=""),
    ])

    assert [op.external_id for op in operations] == ["a"]


def test_foreign_instruments_are_skipped_and_counted():
    operations, warnings = normalize_operations([
        _row("Stock purchase", "BMW.DE", "BMW", amount=-1000.0, oid="a", comment="OPEN BUY 10 @ 100.00"),
        _row("Stock purchase", "KRU.PL", "Kruk", amount=-1000.0, oid="b", comment="OPEN BUY 10 @ 100.00"),
    ])

    assert [op.external_id for op in operations] == ["b"]
    assert len(warnings) == 1
    assert "BMW.DE" in warnings[0]


def test_cash_rows_without_a_ticker_are_kept():
    # Deposits and interest carry no ticker but are still real operations.
    operations, warnings = normalize_operations([
        _row("Deposit", amount=5000.0, oid="a"),
        _row("Free funds interest", amount=0.07, oid="b"),
    ])

    assert [op.external_id for op in operations] == ["a", "b"]
    assert warnings == []


def test_payment_operator_comments_are_discarded_for_cash_movements():
    # A deposit's comment carries an account id with no analytic value; a
    # dividend's comment carries the rate per share and is worth keeping.
    operations, _ = normalize_operations([
        _row("IKZE deposit", amount=6473.97, oid="a",
             comment="Transfer in operation on account with id 52472380"),
        _row("Dividend", "PAS.PL", "Passus", amount=45.0, oid="b",
             comment="PAS.PL PLN 2.5000/ SHR"),
    ])

    by_id = {op.external_id: op for op in operations}
    assert by_id["a"].comment is None
    assert by_id["b"].comment == "PAS.PL PLN 2.5000/ SHR"


def test_timestamps_keep_microsecond_precision():
    # Three PAS payouts in IKZE share a second and differ only in microseconds.
    # Truncating precision would collapse them into one.
    operations, _ = normalize_operations([
        _row("Dividend", "PAS.PL", "Passus", amount=13.74, oid="a", micro=332000),
        _row("Dividend", "PAS.PL", "Passus", amount=45.0, oid="b", micro=278000),
        _row("Dividend", "PAS.PL", "Passus", amount=37.5, oid="c", micro=218000),
    ])

    assert len({op.occurred_at for op in operations}) == 3
    assert operations[0].occurred_at.microsecond == 332000


def test_three_payouts_sharing_a_second_stay_three_dividends():
    # The real IKZE file has exactly this: PAS paid three times at 09:54:01,
    # distinguishable only by microseconds. Any pairing or de-duplication keyed
    # on (ticker, second) would collapse 96.24 zl into 45.00.
    operations, _ = normalize_operations([
        _row("Dividend", "PAS.PL", "Passus", amount=13.74, oid="a", micro=332000),
        _row("Dividend", "PAS.PL", "Passus", amount=45.0, oid="b", micro=278000),
        _row("Dividend", "PAS.PL", "Passus", amount=37.5, oid="c", micro=218000),
    ])

    dividends = extract_dividends(operations)

    assert len(dividends) == 3
    assert sum(d.amount_pln for d in dividends) == pytest.approx(96.24)


def test_dividends_and_their_tax_are_returned_unpaired():
    # Pairing by (ticker, timestamp) fails on real data in 24 cases: the stamps
    # drift by seconds and one payout is sometimes split across rows. Gross and
    # tax stay separate events; the totals are formed later, in SQL.
    operations, _ = normalize_operations([
        _row("Dividend", "KRU.PL", "Kruk", amount=722.0, oid="a"),
        _row("Withholding tax", "KRU.PL", "Kruk", amount=-137.18, oid="b", day=3),
    ])

    dividends = extract_dividends(operations)

    assert len(dividends) == 2
    assert {d.op_type for d in dividends} == {OP_DIVIDEND, OP_WITHHOLDING_TAX}


def test_extract_dividends_ignores_trades_and_cash_movements():
    operations, _ = normalize_operations([
        _row("Stock purchase", "TOA.PL", "Toya", amount=-4128.24, oid="a",
             comment="OPEN BUY 412 @ 10.02"),
        _row("Deposit", amount=5000.0, oid="b"),
        _row("Free funds interest tax", amount=-1.72, oid="c"),
        _row("Dividend", "TOA.PL", "Toya", amount=41.2, oid="d"),
    ])

    assert [d.external_id for d in extract_dividends(operations)] == ["d"]


_GOLDEN_HEADER = ["Type", "Ticker", "Instrument", "Time", "Amount", "ID", "Comment", "Product"]

# One row per shape found in the real exports. Written newest-first, the way XTB
# writes them, so the end-to-end test also proves FIFO sorts by time rather than
# trusting file order. Real exports carry account numbers and must never be
# committed — every fixture is synthesized instead.
_GOLDEN_ROWS = [
    # Foreign instrument: priced in EUR in the comment while Amount is PLN.
    ["Stock purchase", "BMW.DE", "BMW", datetime(2026, 7, 9, 10, 0), -1000.0, "20",
     "OPEN BUY 10 @ 90.00", "My Trades"],
    # Dividend with no tax at all — the IKZE case, where gross equals net.
    ["Dividend", "TOA.PL", "Toya", datetime(2026, 7, 8, 10, 0), 41.2, "19", "TOA.PL PLN 0.1000/ SHR", "IKZE"],
    # Three payouts inside one second, distinguishable only by microseconds.
    ["Dividend", "PAS.PL", "Passus", datetime(2026, 7, 7, 9, 54, 1, 332000), 13.74, "18", "PAS.PL PLN 2.5000/ SHR", "My Trades"],
    ["Dividend", "PAS.PL", "Passus", datetime(2026, 7, 7, 9, 54, 1, 278000), 45.0, "17", "PAS.PL PLN 2.5000/ SHR", "My Trades"],
    ["Dividend", "PAS.PL", "Passus", datetime(2026, 7, 7, 9, 54, 1, 218000), 37.5, "16", "PAS.PL PLN 2.5000/ SHR", "My Trades"],
    # Dividend and its withholding tax, deliberately left unpaired.
    ["Withholding tax", "KRU.PL", "Kruk", datetime(2026, 7, 6, 12, 0), -137.18, "15", "", "My Trades"],
    ["Dividend", "KRU.PL", "Kruk", datetime(2026, 7, 6, 11, 0), 722.0, "14", "KRU.PL PLN 20.0000/ SHR", "My Trades"],
    # A tax that is NOT a dividend tax.
    ["Free funds interest tax", "", "", datetime(2026, 7, 5, 10, 0), -1.72, "13", "", "My Trades"],
    # Ticker sold down to zero.
    ["Stock sell", "PZU.PL", "PZU", datetime(2026, 7, 5, 10, 0), 275.0, "12", "CLOSE BUY 5 @ 55.00", "My Trades"],
    ["Stock sell", "SNT.PL", "Synektik", datetime(2026, 7, 4, 10, 0), 2500.0, "11", "CLOSE BUY 10 @ 250.00", "My Trades"],
    ["Stock purchase", "SNT.PL", "Synektik", datetime(2026, 7, 3, 10, 0), -2000.0, "10", "OPEN BUY 10 @ 200.00", "My Trades"],
    # Fractional shares, the signature of XTB's Investment Plans product.
    ["Stock purchase", "ETFBW20TR.PL", "Beta ETF WIG20TR", datetime(2026, 7, 2, 10, 0), -539.47, "9",
     "OPEN BUY 10.2017 @ 52.89", "Investment Plans"],
    # One order for 11 split into two fill rows, 9 and 2.
    ["Stock purchase", "CBF.PL", "CyberFolks", datetime(2026, 7, 2, 9, 1), -376.8, "8", "OPEN BUY 2/11 @ 188.40", "My Trades"],
    ["Stock purchase", "CBF.PL", "CyberFolks", datetime(2026, 7, 2, 9, 0), -1695.6, "7", "OPEN BUY 9/11 @ 188.40", "My Trades"],
    ["Stock purchase", "PZU.PL", "PZU", datetime(2026, 7, 1, 10, 0), -250.0, "6", "OPEN BUY 5 @ 50.00", "My Trades"],
    ["Stock purchase", "SNT.PL", "Synektik", datetime(2026, 7, 1, 9, 0), -1000.0, "5", "OPEN BUY 10 @ 100.00", "My Trades"],
    ["Deposit", "", "", datetime(2026, 7, 1, 8, 0), 5000.0, "4", "Transfer in operation on account with id 51667579", "My Trades"],
    ["Total", "", "", None, 99999.0, "", "", ""],
]


def _golden_workbook(*, extra_sheet: str | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Operations"
    ws.append(["Account number", "51667579"])
    ws.append(["Cash Operations", ""])
    ws.append(["Date from (UTC)", datetime(2006, 1, 1)])
    ws.append(["Date to (UTC)", datetime(2026, 7, 29)])
    ws.append(_GOLDEN_HEADER)
    for row in _GOLDEN_ROWS:
        ws.append(row)
    if extra_sheet is not None:
        # Real exports carry more than one sheet; the parser indexes none of them.
        other = wb.create_sheet(extra_sheet)
        other.append(["Symbol", "Volume", "Open price"])
        other.append(["CBF", 100, 12.5])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_golden_export_reconstructs_the_expected_positions():
    result = parse_xtb_export(_golden_workbook())

    by_ticker = {p.ticker: p for p in result.positions}
    assert sorted(by_ticker) == ["CBF", "ETFBW20TR", "SNT"]
    # Both fill rows of the same 11-share order counted once each.
    assert by_ticker["CBF"].shares == pytest.approx(11.0)
    assert by_ticker["CBF"].avg_buy_price == pytest.approx(188.40)
    # FIFO: the 100.00 lot was sold, the 200.00 lot remains.
    assert by_ticker["SNT"].shares == pytest.approx(10.0)
    assert by_ticker["SNT"].avg_buy_price == pytest.approx(200.0)
    assert by_ticker["ETFBW20TR"].shares == pytest.approx(10.2017)
    assert by_ticker["ETFBW20TR"].avg_buy_price == pytest.approx(52.89)


def test_golden_export_reports_closed_tickers_and_warnings():
    result = parse_xtb_export(_golden_workbook())

    assert result.closed_tickers == ["PZU"]
    assert len(result.warnings) == 1
    assert "BMW.DE" in result.warnings[0]


def test_golden_export_dividend_totals_match():
    result = parse_xtb_export(_golden_workbook())

    gross = sum(d.amount_pln for d in result.dividends if d.op_type == OP_DIVIDEND)
    tax = sum(d.amount_pln for d in result.dividends if d.op_type == OP_WITHHOLDING_TAX)

    # 722.00 + 13.74 + 45.00 + 37.50 + 41.20 — the interest tax stays out of it.
    assert gross == pytest.approx(859.44)
    assert tax == pytest.approx(-137.18)
    assert len(result.dividends) == 6


def test_golden_export_keeps_every_row_it_did_not_reject():
    result = parse_xtb_export(_golden_workbook())

    # 18 rows in, minus the Total row and the foreign instrument.
    assert len(result.operations) == 16
    assert all(op.external_id for op in result.operations)


def test_external_ids_are_unique_so_the_import_can_dedupe_on_them():
    result = parse_xtb_export(_golden_workbook())

    ids = [op.external_id for op in result.operations]
    assert len(ids) == len(set(ids))


def test_parser_package_imports_no_data_or_web_layer():
    # `tach check` cannot catch this: tach.toml declares `src` with
    # depends_on = [{ path = "db" }], so importing BigQuery from src/ is
    # explicitly allowed and the check stays green either way. The parser being
    # pure is what lets the golden dataset run without any infrastructure.
    forbidden = ("db", "fastapi", "google", "starlette")
    package = pathlib.Path(__file__).resolve().parents[1] / "src" / "brokers"

    offenders = []
    for path in sorted(package.glob("*.py")):
        tree = ast.parse(path.read_text(encoding="utf-8"))
        for node in ast.walk(tree):
            if isinstance(node, ast.Import):
                names = [alias.name for alias in node.names]
            elif isinstance(node, ast.ImportFrom):
                names = [node.module or ""]
            else:
                continue
            for name in names:
                if name.split(".")[0] in forbidden:
                    offenders.append(f"{path.name}: {name}")

    assert offenders == []


def test_registry_exposes_xtb_and_rejects_an_unknown_broker():
    assert [b["id"] for b in list_brokers()] == ["xtb"]
    assert list_brokers()[0]["label"] == "XTB"
    assert get_parser("xtb") is parse_xtb_export

    with pytest.raises(UnknownBrokerError):
        get_parser("mbank")


def test_buys_are_ordered_by_time_not_by_input_order():
    # Cash Operations rows are not guaranteed chronological; FIFO must sort.
    out_of_order = [
        _buy("DIG", 10, 200.0, day=9),
        _buy("DIG", 10, 100.0, day=2),
        _sell("DIG", 10, 250.0, day=10),
    ]

    positions, _ = reconstruct_positions(out_of_order)

    # The day-2 lot at 100 is the oldest and must be the one consumed.
    assert positions[0].avg_buy_price == pytest.approx(200.0)


# ── size ceiling (PUL-105) ────────────────────────────────────────────────────


def test_skipping_unread_sheets_changes_nothing_the_parser_produces():
    # The reader stopped materialising sheets no parser declares columns for.
    # That is a memory change, so it has to be observationally invisible: the
    # same export with an extra sheet must yield the same positions, dividends
    # and cash as without it.
    plain = parse_xtb_export(_golden_workbook())
    with_extra = parse_xtb_export(_golden_workbook(extra_sheet="Open Positions"))

    assert [(p.ticker, p.shares) for p in with_extra.positions] == [
        (p.ticker, p.shares) for p in plain.positions
    ]
    assert with_extra.closed_tickers == plain.closed_tickers
    assert len(with_extra.dividends) == len(plain.dividends)
    assert with_extra.cash_pln == plain.cash_pln


# ── free cash (PUL-95 Phase 8) ────────────────────────────────────────────────


def test_cash_balance_comes_from_the_total_row_not_from_the_parsed_rows():
    """The Total is the broker's own net of every movement — including the ones
    the parser deliberately drops. Re-adding only what was imported gives a
    different number (measured on the real Glowny export: 84.03 vs 143.94, the
    gap being foreign trades)."""
    rows = [
        _row("Stock purchase", "PKO.PL", "PKO", amount=-1000.0, comment="OPEN BUY 10 @ 100.00"),
        _row("Stock purchase", "AAPL.US", "Apple", amount=-500.0, comment="OPEN BUY 2 @ 250.00"),
        _row("Deposit", amount=2000.0, oid="d1"),
        {"Type": "Total", "Ticker": None, "Instrument": None, "Time": None, "Amount": 500.0},
    ]

    assert extract_cash_balance(rows) == pytest.approx(500.0)

    # And the parsed operations really do disagree — the foreign row is dropped.
    operations, warnings = normalize_operations(rows)
    assert sum(op.amount_pln for op in operations) == pytest.approx(1000.0)
    assert len(warnings) == 1


def test_a_missing_total_row_means_unknown_not_zero():
    """Zero cash and unknown cash are different facts and must not render alike."""
    assert extract_cash_balance([_row("Deposit", amount=100.0)]) is None
    assert extract_cash_balance([]) is None


def test_a_non_numeric_total_does_not_take_the_import_down():
    rows = [{"Type": "Total", "Amount": "n/d"}]
    assert extract_cash_balance(rows) is None


def test_parse_export_carries_the_cash_balance_through():
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Cash Operations"
    sheet.append(["Type", "Ticker", "Instrument", "Time", "Amount", "ID", "Comment"])
    sheet.append(["Stock purchase", "PKO.PL", "PKO", datetime(2026, 1, 2, 9, 0),
                  -1000.0, "1", "OPEN BUY 10 @ 100.00"])
    sheet.append(["Deposit", None, None, datetime(2026, 1, 1, 9, 0), 1500.0, "2", ""])
    sheet.append(["Total", None, None, None, 500.0, None, None])
    buffer = io.BytesIO()
    wb.save(buffer)

    result = parse_xtb_export(buffer.getvalue())

    assert result.cash_pln == pytest.approx(500.0)
